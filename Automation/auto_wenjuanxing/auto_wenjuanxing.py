import time
import asyncio
from pyppeteer import launch
from pyppeteer_stealth import stealth
from openpyxl import load_workbook


class AutoWenJuanXing(object):

    def __init__(self, answer_name):
        self.answer_name = answer_name
        self.answer_file_path = "C:\\workspace\\附件1：党史知识学习手册.xlsx"
        self.urls = ["https://ks.wjx.top/vm/Y740Hwy.aspx", "https://ks.wjx.top/vm/mLDDbFL.aspx",
                     "https://ks.wjx.top/vm/PoE33fL.aspx", "https://ks.wjx.top/vm/wcKC5oK.aspx"]

    def pack_answer(self):
        """
        packing answer
        :return: answer package
        """
        wb = load_workbook(self.answer_file_path)
        ws = wb["Table 1"]
        cell_range = ws["H2": "H801"]
        print("检测到共有%s道题" % cell_range.__len__())
        answers = [[cell[0].value.strip() for cell in cell_range][page * 200: (page * 200) + 200] for page in
                   range(self.urls.__len__())]
        q_a_pack = zip(self.urls, answers)
        return q_a_pack

    async def answer(self, pack):
        """
        answer main func
        :param pack: the package of url link and the page answers
        :return:
        """
        browser = await launch({
            # 路径就是你的谷歌浏览器的安装路径
            'executablePath': 'C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe',
            # Pyppeteer 默认使用的是无头浏览器,所以要显示需要给False
            'headless': False,
            # 设置Windows-size和Viewport大小来实现网页完整显示
            'args': ['--no-sandbox', '--window-size=1366,850']
        })
        for url, page_answer in pack:
            # 创建页面
            page = await browser.newPage()

            # 设置页面属性
            await page.setViewport({'width': 1366, 'height': 768})

            # 反检测关键
            await stealth(page)

            # 关闭多余标签页
            for _page in await browser.pages():
                if _page != page:
                    await _page.close()

            # 开始访问url答题
            await page.goto(url)

            # 姓名
            await page.type('#q1', self.answer_name)

            # 单选、多选
            for index, question_answers in enumerate(page_answer):
                ans_dict = {"A": 1, "B": 2, "C": 3, "D": 4}
                for answer in question_answers:
                    xpath = """//div[@id="div%s"]/div[@class="ui-controlgroup"]/div[position()=%s]""" % (
                    index + 2, ans_dict[answer])
                    print(index + 1, answer, xpath)
                    button = await page.xpath(xpath)
                    await button[0].click()

            # 延迟2秒提交
            time.sleep(2)

            # 提交
            submit_btn = await page.xpath("""//div[@id="ctlNext"]""")
            await submit_btn[0].click()

            # 提交完毕跳转延迟5秒
            time.sleep(5)

    def run(self):
        q_a_pack = self.pack_answer()
        asyncio.get_event_loop().run_until_complete(self.answer(q_a_pack))


if __name__ == "__main__":
    AutoWenJuanXing('姓名').run()

# selenium版本，会碰上问卷星的检测

# from selenium import webdriver
# driver = webdriver.Chrome("C:\\LanguageEnvs\\Anaconda3\\Lib\\site-packages\\selenium\\webdriver\\chrome\\chromedriver.exe")
#
# driver.get("https://ks.wjx.top/vm/Y740Hwy.aspx")
# driver.find_element_by_xpath("""//input[@id='q1']""").send_keys("李相廷")
#
# for i in range(1, 1001):
#     question_answers = answers[i-1]
#     ans_dict = {"A": 1, "B": 2, "C": 3, "D" : 4}
#     for answer in question_answers:
#         xpath = """//div[@id="div%s"]/div[@class="ui-controlgroup"]/div[position()=%s]"""%(i+1, ans_dict[answer])
#         print(i, answer, xpath)
#         driver.find_element_by_xpath(xpath).click()
#     print("%s题答题完成"%i)
